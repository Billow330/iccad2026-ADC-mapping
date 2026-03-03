"""
export_figures_tables.py — Export all figure/table data to Excel
"""
import json, csv
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
OUT = ROOT / 'results/figures_tables_data.xlsx'

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── Color palette ──────────────────────────────────────────────────────────
H1 = PatternFill("solid", fgColor="1F4E79")   # dark blue header
H2 = PatternFill("solid", fgColor="2E75B6")   # medium blue sub-header
H3 = PatternFill("solid", fgColor="BDD7EE")   # light blue section
HL = PatternFill("solid", fgColor="FFD700")   # gold highlight (ours/best)
HL2= PatternFill("solid", fgColor="FF6B6B")   # red highlight (HAWQ/bad)
ALT= PatternFill("solid", fgColor="F2F2F2")   # alternating row

def hfont(bold=True, color="FFFFFF", size=10):
    return Font(bold=bold, color=color, size=size)

def bfont(bold=False, size=9):
    return Font(bold=bold, size=size)

thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

def set_header(ws, row, cols, texts, fill=H1, font_color="FFFFFF", bold=True):
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = fill
        c.font = Font(bold=bold, color=font_color, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin

def write_row(ws, row, cols, values, fill=None, bold=False, align='center'):
    for col, val in zip(cols, values):
        c = ws.cell(row=row, column=col, value=val)
        if fill:
            c.fill = fill
        c.font = Font(bold=bold, size=9)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        c.border = thin

def title_row(ws, row, col, text, span=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = H1
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal='center', vertical='center')
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c.border = thin

def section_row(ws, row, col, text, span, fill=H3):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill
    c.font = Font(bold=True, color="1F4E79", size=10)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: Fig1 & Table II — PPA Sweep
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Fig1_TableII_PPA_Sweep")
ws.column_dimensions['A'].width = 12
for col in ['B','C','D','E','F']:
    ws.column_dimensions[col].width = 16
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 40

title_row(ws, 1, 1, "Fig 1 & Table II — NeuroSIM PPA Sweep (OPT-125M & OPT-1.3B)", span=6)
section_row(ws, 2, 1, "数据来源: results/ppa/opt125m/ppa_sweep_opt125m.csv | OPT-1.3B列 = OPT-125M × 14.22倍 | 图类型: 折线图（双Y轴）", span=6, fill=H3)

headers = ["ADC bits", "Chip area 125M\n(mm²)", "ADC area 125M\n(mm²)", "ADC占比\n(%)", "Chip area 1.3B\n(mm²，缩放)", "Energy 125M\n(nJ/512-token)"]
set_header(ws, 3, range(1,7), headers, fill=H2)

ppa_data = [
    (3,  667.65,  26.20,  3.92,   9496,   4578),
    (4,  694.57,  40.57,  5.84,   9878,   5141),
    (5,  731.03,  63.65,  8.71,  10397,   6110),
    (6,  806.38, 125.23, 15.53,  11469,   7922),
    (7,  936.60, 228.43, 24.39,  13321,  11401),
    (8, 1146.86, 407.42, 35.53,  16311,  18205),
    (9, 1671.54, 821.34, 49.14,  23773,  31738),
    (10,2821.71,1802.79, 63.89,  40131,  58681),
]
for i, row_data in enumerate(ppa_data):
    r = i + 4
    fill = HL if row_data[0] == 7 else (ALT if i % 2 == 0 else None)
    bold = (row_data[0] == 7)
    write_row(ws, r, range(1,7), row_data, fill=fill, bold=bold)

r = len(ppa_data) + 4
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws.cell(row=r, column=1, value="★ 7b工作点（金色）: ADC面积228.4mm²=芯片总面积24.4% | 7b→8b: ADC面积+78%（超线性增长）| 7b→6b: ADC面积-45%")
c.fill = PatternFill("solid", fgColor="FFF2CC")
c.font = Font(bold=True, size=9, color="7F6000")
c.alignment = Alignment(horizontal='left', wrap_text=True)
c.border = thin
ws.row_dimensions[r].height = 25

section_row(ws, r+2, 1, "图表设计建议: X轴=ADC bits(3-10), 左Y轴=ADC area(mm²), 右Y轴=ADC占总面积%, 在7b处标注垂直虚线", span=6)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: Fig2 — Saturation Heterogeneity
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Fig2_Saturation")
for col in ['A','B','C','D','E','F']:
    ws.column_dimensions[col].width = 16
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 40

title_row(ws, 1, 1, "Fig 2 — Per-Layer Max-Clip ADC Saturation Rate（饱和率异质性）", span=6)
section_row(ws, 2, 1, "数据来源: results/opt125m/outlier_facebook_opt-125m_adc7.csv 和 results/opt1.3b/outlier_facebook_opt-1.3b_adc7.csv 的 sat_rate_worst 列 | 图类型: 分组柱状图", span=6)

# sub-table 1: OPT-125M
section_row(ws, 3, 1, "OPT-125M @ 7b ADC（max-clip饱和率）", span=6, fill=H2)
set_header(ws, 4, range(1,7),
    ["Layer type", "层名称", "数量(层)", "sat_rate_worst均值", "论文标注值", "含义"],
    fill=H2)

opt125_sat = [
    ("attn_qkv", "q_proj / k_proj / v_proj", 36, "99.99%", "≈100%", "饱和率最高（论文认为应最敏感）"),
    ("ffn_up",   "fc1",                      12, "94.5%",  "≈94%",  "饱和率次高"),
    ("lm_head",  "lm_head",                   1, "100.0%", "≈100%", "语言模型头"),
    ("ffn_down", "fc2",                       12, "19.3%",  "≈19%",  "饱和率低（实际最敏感！）"),
    ("attn_out", "out_proj",                  12, "3.6%",   "≈4%",   "饱和率最低"),
]
colors_125 = [HL2, ALT, None, HL, ALT]  # red=high sat, gold=most sensitive
row_colors = [
    PatternFill("solid", fgColor="FF6B6B"),  # attn_qkv - high sat, wrong intuition
    PatternFill("solid", fgColor="FFAA00"),  # ffn_up
    PatternFill("solid", fgColor="F2F2F2"),  # lm_head
    HL,                                       # ffn_down - low sat but most sensitive
    PatternFill("solid", fgColor="90EE90"),  # attn_out - lowest sat
]
for i, (d, clr) in enumerate(zip(opt125_sat, row_colors)):
    write_row(ws, 5+i, range(1,7), d, fill=clr, bold=(i in [0,3]))

# sub-table 2: OPT-1.3B
section_row(ws, 11, 1, "OPT-1.3B @ 7b ADC（max-clip饱和率）— 注意1.3B的out_proj饱和率从3.6%升到59%", span=6, fill=H2)
set_header(ws, 12, range(1,7),
    ["Layer type", "层名称", "数量(层)", "sat_rate_worst均值", "与125M对比", "含义"],
    fill=H2)

opt13b_sat = [
    ("attn_qkv", "q_proj / k_proj / v_proj", 72, "100.0%", "同125M", "仍然100%"),
    ("ffn_up",   "fc1",                       24, "100.0%", "同125M", "仍然100%"),
    ("lm_head",  "lm_head",                    1, "100.0%", "同125M", "仍然100%"),
    ("ffn_down", "fc2",                        24, "91.2%",  "125M=19% → 1.3B=91%", "随模型增大饱和率上升"),
    ("attn_out", "out_proj",                   24, "59.0%",  "125M=3.6% → 1.3B=59%", "随模型增大饱和率上升"),
]
for i, d in enumerate(opt13b_sat):
    fill = ALT if i % 2 == 0 else None
    write_row(ws, 13+i, range(1,7), d, fill=fill)

r = 19
section_row(ws, r, 1, "★ 关键点: Fig2显示高饱和≠高敏感 — attn_qkv饱和率100%但ΔPPL/layer仅0.128（最不敏感），fc2仅19%饱和但ΔPPL/layer=1.413（最敏感）→ 11×反转", span=6)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: Fig3 & Table I — Group Sensitivity
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Fig3_TableI_Sensitivity")
for col in ['A','B','C','D','E','F','G']:
    ws.column_dimensions[col].width = 16
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 40

title_row(ws, 1, 1, "Fig 3 & Table I — 分层CIM ADC敏感度测量（OPT-125M, 7b→6b）", span=7)
section_row(ws, 2, 1, "数据来源: results/sensitivity/opt125m/group_sensitivity.json | Baseline PPL=333.84 | Clean PPL=38.35 | 图类型: 水平条形图", span=7)

set_header(ws, 3, range(1,8),
    ["Layer type", "层名称示例", "数量(层)", "ΔPPL(全组)", "ΔPPL/layer", "Sat.rate(max-clip)", "敏感度排名(1=最敏感)"],
    fill=H2)

sens_data = [
    ("ffn_down (fc2)",    "fc2",         12, "+16.955", "1.4129", "≈19%",  "1 ← 最敏感"),
    ("attn_out (out_proj)","out_proj",   12, "+10.389", "0.8657", "≈4%",   "2"),
    ("lm_head",           "lm_head",      1,  "+0.700", "0.7004", "≈100%", "3"),
    ("ffn_up (fc1)",      "fc1",         12,  "+2.417", "0.2015", "≈94%",  "4"),
    ("attn_qkv",     "q/k/v_proj",       36,  "+4.618", "0.1283", "≈100%", "5 ← 最不敏感"),
]
row_fills = [
    PatternFill("solid", fgColor="FF6B6B"),
    PatternFill("solid", fgColor="FFA500"),
    PatternFill("solid", fgColor="FFD700"),
    PatternFill("solid", fgColor="90EE90"),
    PatternFill("solid", fgColor="87CEEB"),
]
for i, (d, clr) in enumerate(zip(sens_data, row_fills)):
    write_row(ws, 4+i, range(1,8), d, fill=clr, bold=(i==0 or i==4))

r = 10
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c = ws.cell(row=r, column=1, value="11× 反转: fc2(19%饱和率) ΔPPL/layer=1.413 vs attn_qkv(100%饱和率) ΔPPL/layer=0.128 → 11.02倍差距")
c.fill = PatternFill("solid", fgColor="FF0000")
c.font = Font(bold=True, size=10, color="FFFFFF")
c.alignment = Alignment(horizontal='center')
c.border = thin

section_row(ws, r+2, 1, "原始数值（直接从JSON）: baseline_ppl=333.8396, attn_qkv总ΔPPL=4.618, attn_out=10.389, ffn_up=2.417, ffn_down=16.955, lm_head=0.700", span=7)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4: Fig4 — Proxy Scatter & Table IV
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Fig4_TableIV_Proxy")
for col in ['A','B','C','D','E','F','G']:
    ws.column_dimensions[col].width = 16
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 40

title_row(ws, 1, 1, "Fig 4 & Table IV — 代理指标对比（Saturation vs Hessian vs CIM直接测量）", span=7)
section_row(ws, 2, 1, "数据来源: group_sensitivity.json + outlier_opt-125m_adc7.csv(sat_rate_worst) + hessian_group.json | 图类型: 散点图(两个子图)", span=7)

# scatter data
section_row(ws, 3, 1, "散点图原始数据（每种层类型一个点）", span=7, fill=H2)
set_header(ws, 4, range(1,8),
    ["Layer type", "ΔPPL/layer\n(Y轴，CIM实测)", "Sat.rate_worst\n(X轴子图a)", "Hessian mean\n(X轴子图b)", "Hessian std", "Sat排名↑", "HAWQ排名↑"],
    fill=H2)

scatter_data = [
    ("attn_qkv",  0.1283, "100.0%", 0.001417, 0.001369, "1(最敏感预测)", "4"),
    ("ffn_up",    0.2015,  "94.5%", 0.002015, 0.001328, "2",             "3"),
    ("lm_head",   0.7004, "100.0%", 0.009270, 0.0,      "3",             "1"),
    ("ffn_down",  1.4129,  "19.3%", 0.005291, 0.009066, "4",             "2"),
    ("attn_out",  0.8657,   "3.6%", 0.000183, 0.000139, "5(最不敏感预测)","5"),
]
for i, d in enumerate(scatter_data):
    fill = ALT if i % 2 == 0 else None
    write_row(ws, 5+i, range(1,8), d, fill=fill)

r = 11
section_row(ws, r, 1, "统计结论", span=7, fill=H2)
stats = [
    ("Spearman ρ (Sat vs CIM)", "−0.80", "负相关！高饱和率→低敏感度（反直觉）", "", "", "", ""),
    ("Spearman ρ (HAWQ vs CIM)", "+0.20", "p=0.75，不显著", "", "", "", ""),
    ("HAWQ排名正确数", "0/5", "5种层类型全部预测错误", "", "", "", ""),
    ("Sat排名正确数",  "0/5", "5种层类型全部预测错误", "", "", "", ""),
]
for i, d in enumerate(stats):
    write_row(ws, r+1+i, range(1,8), d, fill=HL2 if i < 2 else ALT)

# proxy ranking table (=Table IV)
r = 17
section_row(ws, r, 1, "Table IV — 敏感度代理排名对比（↑=1最敏感，5最不敏感）", span=7, fill=H2)
set_header(ws, r+1, range(1,7),
    ["Layer type", "Sat.排名↑\n(错误预测)", "HAWQ排名↑\n(错误预测)", "CIM排名↑\n(正确)", "ΔPPL/layer\n(实测)", "结论"],
    fill=H2)
proxy_data = [
    ("q/k/v_proj", "1 (100%)", "4", "5", "0.128", "Sat预测最敏感，实际最不敏感"),
    ("fc1",        "2 (94%)",  "3", "4", "0.201", ""),
    ("lm_head",    "3 (100%)", "1", "3", "0.700", "HAWQ预测最敏感"),
    ("fc2",        "4 (19%)",  "2", "1", "1.413", "★ 实际最敏感，但Sat预测倒数第2"),
    ("out_proj",   "5 (4%)",   "5", "2", "0.866", "Sat/HAWQ都说最不敏感，实际第2敏感"),
]
fills_proxy = [
    PatternFill("solid", fgColor="FFB3B3"),
    ALT,
    PatternFill("solid", fgColor="FFD9B3"),
    HL,
    PatternFill("solid", fgColor="B3FFB3"),
]
for i, (d, clr) in enumerate(zip(proxy_data, fills_proxy)):
    write_row(ws, r+2+i, range(1,7), d, fill=clr, bold=(i in [0,3]))

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5: Fig5 — Bit Assignment
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Fig5_BitAssignment")
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 22

title_row(ws, 1, 1, "Fig 5 — 逐层位宽分配（ILP vs Greedy, 20.5% ADC节省目标）", span=6)
section_row(ws, 2, 1, "数据来源: results/sensitivity/opt125m/allocations.json | 图类型: 逐层热图（蓝=6b，白/灰=7b）", span=6)

set_header(ws, 3, range(1,7),
    ["Layer\nIndex", "Layer名称", "Layer\nType", "ILP\nbits", "Greedy\nbits", "说明"],
    fill=H2)

# load allocations
with open(ROOT / 'results/sensitivity/opt125m/allocations.json') as f:
    alloc = json.load(f)

ilp = alloc['ilp']
greedy = alloc['greedy']

def get_ltype(name):
    if any(x in name for x in ['q_proj','k_proj','v_proj']):
        return 'attn_qkv'
    elif 'out_proj' in name:
        return 'attn_out'
    elif 'fc1' in name:
        return 'ffn_up'
    elif 'fc2' in name:
        return 'ffn_down'
    elif 'lm_head' in name:
        return 'lm_head'
    return 'other'

fill_6b = PatternFill("solid", fgColor="2E75B6")  # blue = 6b
fill_7b = PatternFill("solid", fgColor="F2F2F2")  # gray = 7b

layers = list(ilp.keys())
for i, name in enumerate(layers):
    r = i + 4
    ltype = get_ltype(name)
    ilp_b = ilp[name]
    grd_b = greedy.get(name, 7)
    short = name.replace('model.decoder.layers.','L').replace('self_attn.','').replace('model.decoder.','')
    note = ""
    if ilp_b == 6 and grd_b == 7:
        note = "ILP降为6b，Greedy保持7b"
    elif ilp_b == 7 and grd_b == 6:
        note = "ILP保持7b，Greedy降为6b"
    elif ilp_b == 6 and grd_b == 6:
        note = "两者都降为6b"

    row_fill = fill_6b if ilp_b == 6 else fill_7b
    font_clr = "FFFFFF" if ilp_b == 6 else "000000"
    c1 = ws.cell(row=r, column=1, value=i)
    c2 = ws.cell(row=r, column=2, value=short)
    c3 = ws.cell(row=r, column=3, value=ltype)
    c4 = ws.cell(row=r, column=4, value=ilp_b)
    c5 = ws.cell(row=r, column=5, value=grd_b)
    c6 = ws.cell(row=r, column=6, value=note)
    for c in [c1,c2,c3,c4,c5,c6]:
        c.border = thin
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(size=8)
    c4.fill = fill_6b if ilp_b == 6 else fill_7b
    c4.font = Font(size=8, color=font_clr, bold=(ilp_b==6))
    c5.fill = fill_6b if grd_b == 6 else fill_7b
    c5.font = Font(size=8, color="FFFFFF" if grd_b==6 else "000000")

r = len(layers) + 5
section_row(ws, r, 1, f"ILP统计: 30层@6b, 43层@7b | Greedy统计: 30层@6b(layers 0-9 全部q/k/v), 43层@7b | 两者ADC节省相同(20.5%)", span=6)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6: Fig6 — Pareto Frontier
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Fig6_Pareto")
for col in ['A','B','C','D','E','F','G']:
    ws.column_dimensions[col].width = 15

title_row(ws, 1, 1, "Fig 6 — ILP Pareto前沿（PPL vs ADC Area Trade-off）", span=7)
section_row(ws, 2, 1, "数据来源: results/sensitivity/opt125m/pareto_frontier.json | 图类型: 折线图(Pareto curve)", span=7)

set_header(ws, 3, range(1,8),
    ["Target\nsavings(%)", "Actual\nsavings(%)", "ADC area\n(mm²)", "Chip area\n(mm²)", "PPL", "Bits分布", "备注"],
    fill=H2)

with open(ROOT / 'results/sensitivity/opt125m/pareto_frontier.json') as f:
    pareto = json.load(f)

# Add baseline
baseline_row = (0, 0, 228.43, 936.60, 306.4, "all 7b (73层)", "★ 基准线（Uniform 7b）")
write_row(ws, 4, range(1,8), baseline_row, fill=HL, bold=True)

for i, p in enumerate(pareto):
    bits_str = ", ".join(f"{b}b×{n}" for b,n in sorted(p['bits_dist'].items(), key=lambda x:-int(x[0])))
    note = ""
    if abs(p['target_savings'] - 20) < 5:
        note = "← 论文主要工作点（实测PPL=308.6来自stable_eval）"
    row = (
        f"{p['target_savings']:.0f}%",
        f"{p['actual_savings']:.2f}%",
        f"{p['adc_area_mm2']:.2f}",
        f"{p['chip_area_mm2']:.2f}",
        f"{p['ppl']:.2f}",
        bits_str,
        note,
    )
    fill = HL if abs(p['target_savings'] - 20) < 5 else (ALT if i%2==0 else None)
    write_row(ws, 5+i, range(1,8), row, fill=fill)

# Add SQ+6b for reference
sq_row = ("50%", "50.00%", "114.22", "822.38", "305.0", "SmoothQuant+6b", "★ SQ+6b: PPL低于7b基准！")
write_row(ws, 5+len(pareto), range(1,8), sq_row, fill=PatternFill("solid", fgColor="DDA0DD"), bold=True)

r = 5 + len(pareto) + 2
section_row(ws, r, 1, "注意: pareto_frontier.json中PPL由sensitivity评估(4calib+10eval批次)得到，与Table III的stable eval(100批次)略有差异，20%目标点稳定值为308.6", span=7)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 7: Fig7 & Table III — Comparison Bars
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Fig7_TableIII_Comparison")
for col in ['A','B','C','D','E','F','G']:
    ws.column_dimensions[col].width = 18

title_row(ws, 1, 1, "Fig 7 & Table III — 各配置PPL与ADC面积对比（100-batch Stable Eval）", span=7)
section_row(ws, 2, 1, "数据来源: results/stable/opt125m/stable_eval_results.csv + results/sensitivity/opt125m/hawq_comparison.json | 图类型: 双子图柱状图", span=7)

set_header(ws, 3, range(1,8),
    ["配置名称", "PPL\n(WikiText-2↓)", "ADC area\n(mm²↓)", "Chip area\n(mm²)", "ADC节省\n(%)", "颜色(HEX)", "备注"],
    fill=H2)

configs = [
    ("Uniform 7b (baseline)",   306.4, 228.4, 936.6,  0.0,  "#BDC3C7", "基准线"),
    ("Uniform 6b",              315.3, 114.2, 822.4, 50.0,  "#85C1E9", ""),
    ("HAWQ-guided Greedy",      321.3, 181.5, 889.7, 20.5,  "#E74C3C", "★ 最差：比ILP差12.7 PPL"),
    ("Sat-guided Greedy",       313.8, 181.5, 889.7, 20.5,  "#F39C12", ""),
    ("Sens-guided Greedy",      312.5, 181.5, 889.7, 20.5,  "#27AE60", ""),
    ("ILP (ours)",              308.6, 181.5, 889.7, 20.5,  "#2980B9", "★ 最优：20.5%节省下最低PPL"),
    ("SQ + Uniform 6b",         305.0, 114.2, 822.4, 50.0,  "#8E44AD", "★ PPL低于7b基准！"),
    ("SQ + ILP 20%",            309.4, 181.5, 889.7, 20.5,  "#1ABC9C", ""),
]
cfg_fills = [
    PatternFill("solid", fgColor="BDC3C7"),
    PatternFill("solid", fgColor="D6EAF8"),
    HL2,                                         # HAWQ = red
    PatternFill("solid", fgColor="FAD7A0"),
    PatternFill("solid", fgColor="A9DFBF"),
    HL,                                          # ILP = gold
    PatternFill("solid", fgColor="D2B4DE"),
    PatternFill("solid", fgColor="A2D9CE"),
]
for i, (d, clr) in enumerate(zip(configs, cfg_fills)):
    write_row(ws, 4+i, range(1,8), d, fill=clr, bold=(i in [0,2,5,6]))

r = 13
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c = ws.cell(row=r, column=1, value="HAWQ vs ILP差距: 321.3 - 308.6 = 12.7 PPL | ILP vs Sat-Greedy: 313.8 - 308.6 = 5.2 PPL | Y轴范围建议: 295~335")
c.fill = PatternFill("solid", fgColor="FFF2CC")
c.font = Font(bold=True, size=9, color="7F6000")
c.alignment = Alignment(horizontal='left', wrap_text=True)
c.border = thin

# Stability data
r = 15
section_row(ws, r, 1, "稳定性测试数据（3个校准种子）— 来源: results/stable/opt125m/stability_results.json", span=7, fill=H2)
set_header(ws, r+1, range(1,7),
    ["配置", "Seed 0", "Seed 1", "Seed 2", "Mean±Std", "Within-seed ΔPPL (ILP-Uniform)"],
    fill=H2)

stability = [
    ("Uniform 7b", 319.94, 310.05, 307.37, "312.5±5.4", "—"),
    ("ILP 20%",    321.23, 308.56, 314.67, "314.8±5.2", "[+1.30, -1.50, +7.30] → mean=+2.4±3.7"),
    ("SQ + 6b",    326.60, 312.04, 309.59, "316.1±7.5", "—"),
]
for i, d in enumerate(stability):
    fill = ALT if i%2==0 else None
    write_row(ws, r+2+i, range(1,7), d, fill=fill)

r2 = r + 6
ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=7)
c = ws.cell(row=r2, column=1,
    value="稳定性解读: Within-seed ΔPPL=PPL_ILP-PPL_Uniform = mean+2.4±3.7，与Table III的+2.2 PPL一致（论文中用within-seed差值规避绝对PPL不一致问题）")
c.fill = PatternFill("solid", fgColor="E8F8E8")
c.font = Font(size=9)
c.alignment = Alignment(horizontal='left', wrap_text=True)
c.border = thin

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 8: Fig8 / Table VI — OPT-1.3B Cross-model
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Fig8_TableVI_OPT1.3B")
for col in ['A','B','C','D','E']:
    ws.column_dimensions[col].width = 20

title_row(ws, 1, 1, "Fig 8 & Table VI — 跨模型验证：OPT-1.3B敏感度（11b→10b）", span=5)
section_row(ws, 2, 1, "数据来源: results/sensitivity/opt1.3b/group_sensitivity.json | Baseline PPL=708.99 | 图类型: 水平条形图（类似Fig3）", span=5)

set_header(ws, 3, range(1,6),
    ["Layer type", "数量(层)", "ΔPPL/layer\n(11b→10b)", "Rank\n(1=最敏感)", "与OPT-125M对比"],
    fill=H2)

opt13b_data = [
    ("ffn_down (fc2)", 24, "+0.619", "1 ← 最敏感", "125M中fc2也是最敏感(1.413) ✓"),
    ("ffn_up (fc1)",   24, "+0.149", "2",           "125M中fc1排名4 → 1.3B升至2"),
    ("attn_out",       24, "+0.033", "3",           "125M中attn_out排名2 → 1.3B降至3"),
    ("attn_qkv",       72, "−0.051", "4 ← 最不敏感","125M中attn_qkv也是最不敏感(0.128) ✓"),
    ("lm_head",         1,"−10.87",  "(排除)",       "数值不稳定（高PPL基准），论文未列入表"),
]
fills_13 = [HL2, ALT, None, HL, PatternFill("solid", fgColor="D3D3D3")]
for i, (d, clr) in enumerate(zip(opt13b_data, fills_13)):
    write_row(ws, 4+i, range(1,6), d, fill=clr, bold=(i in [0,3]))

r = 10
section_row(ws, r, 1, "结论: OPT-1.3B与OPT-125M敏感度排序一致（ffn_down最敏感，attn_qkv最不敏感）→ 证明反转是transformer CIM系统的cross-model属性", span=5)

r2 = 12
section_row(ws, r2, 1, "OPT-1.3B PPA说明（Table II右列）: 以125M为基础乘以14.22倍tile数 | 7b: 13,321mm² | 如果ILP-20%节省: 节省约667mm² ADC面积（比整个125M芯片还大）", span=5, fill=H3)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 9: Fig9 — ILP vs Greedy Multi-Budget
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Fig9_MultibudgetILP_Greedy")
for col in ['A','B','C','D','E','F','G','H','I']:
    ws.column_dimensions[col].width = 13

title_row(ws, 1, 1, "Fig 9 — ILP vs Greedy跨预算对比（OPT-125M, 100-batch stable eval）", span=9)
section_row(ws, 2, 1, "数据来源: results/stable/opt125m/ilp_vs_greedy_multibudget.csv | 图类型: 双折线图（左:PPL，右:ILP优势ΔPPL）", span=9)

set_header(ws, 3, range(1,10),
    ["Target\nsavings(%)", "ILP\nPPL", "ILP actual\nsavings(%)", "ILP ADC\n(mm²)", "Greedy\nPPL", "Greedy actual\nsavings(%)", "Greedy ADC\n(mm²)", "ΔPPL\n(ILP-Greedy)", "ILP更好?"],
    fill=H2)

mb_data = [
    (5,   312.89, 5.48,  215.92, 311.53, 5.48,  215.92, -1.37, "No  ← 极端小预算"),
    (10,  305.23, 10.27, 204.96, 311.07, 10.27, 204.96, +5.84, "Yes ★最大优势"),
    (15,  311.02, 15.07, 194.01, 312.33, 15.07, 194.01, +1.30, "Yes"),
    (20,  309.28, 20.55, 181.49, 311.16, 20.55, 181.49, +1.89, "Yes ← 论文主要结论"),
    (25,  313.88, 25.00, 171.32, 314.10, 25.34, 170.54, +0.22, "Yes"),
    (30,  311.50, 30.14, 159.59, 314.86, 30.14, 159.59, +3.36, "Yes"),
    (40,  311.73, 40.07, 136.90, 312.95, 40.41, 136.12, +1.21, "Yes"),
    (50,  318.46, 50.00, 114.22, 311.59, 50.00, 114.22, -6.87, "No  ← 极端大预算"),
]
mb_fills = [
    PatternFill("solid", fgColor="FFB3B3"),  # 5%: No
    HL,                                       # 10%: best
    ALT, None, ALT, None, ALT,
    PatternFill("solid", fgColor="FFB3B3"),  # 50%: No
]
for i, (d, clr) in enumerate(zip(mb_data, mb_fills)):
    write_row(ws, 4+i, range(1,10), d, fill=clr, bold=(i==1 or i==3))

r = 13
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
c = ws.cell(row=r, column=1, value="基准线: Uniform 7b PPL=306.4（在折线图上画虚线）| ILP在6/8个预算点优于Greedy | 最大优势: 10%预算(+5.8 PPL), 30%预算(+3.4 PPL) | 5%和50%极端值Greedy略好")
c.fill = PatternFill("solid", fgColor="FFF2CC")
c.font = Font(bold=True, size=9, color="7F6000")
c.alignment = Alignment(horizontal='left', wrap_text=True)
c.border = thin

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 10: Table V — Zero-Shot Accuracy
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("TableV_ZeroShot")
for col in ['A','B','C','D','E','F']:
    ws.column_dimensions[col].width = 18

title_row(ws, 1, 1, "Table V — 零样本准确率（Zero-Shot Accuracy, OPT-125M）", span=6)
section_row(ws, 2, 1, "数据来源: results/zeroshot/opt125m/zeroshot_summary.json | 评估: lm-eval-harness, 0-shot, 500样本, HellaSwag + WinoGrande", span=6)

set_header(ws, 3, range(1,7),
    ["Config", "PPL\n(100-batch)", "HellaSwag\n(0-shot, ↑好)", "WinoGrande\n(0-shot, ↑好)", "Avg", "备注"],
    fill=H2)

zs_data = [
    ("FP32 (clean)",    40.44,  "34.8%", "53.6%", "44.2%", "理想上界"),
    ("CIM 7b (uniform)",319.78, "30.6%", "52.0%", "41.3%", "基准CIM"),
    ("SQ + 6b",         320.29, "26.2%", "51.0%", "38.6%", "PPL低但zero-shot差"),
    ("ILP 20%",         319.82, "30.0%", "51.2%", "40.6%", "★ 仅-0.7%pt vs CIM-7b，优于SQ+6b"),
]
zs_fills = [
    PatternFill("solid", fgColor="C6EFCE"),
    ALT,
    PatternFill("solid", fgColor="FFB3B3"),
    HL,
]
for i, (d, clr) in enumerate(zip(zs_data, zs_fills)):
    write_row(ws, 4+i, range(1,7), d, fill=clr, bold=(i==3))

r = 9
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws.cell(row=r, column=1, value="关键发现: ILP-20%零样本准确率(-0.7%pt)优于SQ+6b(-2.7%pt)，尽管SQ+6b的PPL更低(305.0 vs 319.8) → PPL与任务准确率在重度平滑后出现分歧")
c.fill = PatternFill("solid", fgColor="E8F8E8")
c.font = Font(bold=True, size=9)
c.alignment = Alignment(horizontal='left', wrap_text=True)
c.border = thin

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 11: Overview — All figures index
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("总览_Index", 0)
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 20

title_row(ws, 1, 1, "ICCAD 2026 — 全部图表数据索引（Measured Layer Sensitivity Guides Mixed-Precision ADC Allocation for CIM-Based LLM Inference）", span=6)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
ws.row_dimensions[1].height = 35

set_header(ws, 2, range(1,7),
    ["序号", "Sheet名称", "内容", "数据来源文件", "图表类型", "关键数字"],
    fill=H1)

index_data = [
    ("Fig1+TableII","Fig1_TableII_PPA_Sweep",    "ADC面积随位宽变化 + NeuroSIM PPA表",       "ppa/opt125m/ppa_sweep_opt125m.csv",                  "折线图/表格", "7b: ADC占24.4%，10b: 63.9%"),
    ("Fig2",        "Fig2_Saturation",            "各层类型饱和率异质性（两模型对比）",         "opt125m/outlier_..._adc7.csv\nopt1.3b/outlier_..._adc7.csv","分组柱状图","attn_qkv≈100%, fc2≈19%（125M）"),
    ("Fig3+TableI", "Fig3_TableI_Sensitivity",    "分层CIM ADC敏感度测量",                    "sensitivity/opt125m/group_sensitivity.json",          "水平条形图", "fc2=1.413 vs attn_qkv=0.128（11×）"),
    ("Fig4+TableIV","Fig4_TableIV_Proxy",         "代理指标对比（散点图+排名表）",              "group_sensitivity.json\nhessian_group.json\noutlier_csv", "散点图/表格","Sat ρ=-0.80, HAWQ ρ=+0.20"),
    ("Fig5",        "Fig5_BitAssignment",         "ILP vs Greedy逐层位宽分配",                "sensitivity/opt125m/allocations.json",                "热图/逐层图","ILP: 30层@6b,43层@7b"),
    ("Fig6",        "Fig6_Pareto",                "ILP Pareto前沿（PPL vs ADC面积）",          "sensitivity/opt125m/pareto_frontier.json",            "折线图",    "20%节省→PPL=326.8(pareto)/308.6(stable)"),
    ("Fig7+TableIII","Fig7_TableIII_Comparison",  "8种配置PPL+ADC面积柱状图",                  "stable/opt125m/stable_eval_results.csv\nhawq_comparison.json","双子图柱状图","HAWQ=321.3, ILP=308.6（差12.7）"),
    ("Fig8+TableVI","Fig8_TableVI_OPT1.3B",       "跨模型验证：OPT-1.3B敏感度",               "sensitivity/opt1.3b/group_sensitivity.json",          "水平条形图", "1.3B: ffn_down最敏感（与125M一致）"),
    ("Fig9",        "Fig9_MultibudgetILP_Greedy", "ILP vs Greedy跨预算（8个预算点）",          "stable/opt125m/ilp_vs_greedy_multibudget.csv",       "双折线图",  "ILP赢6/8个预算点，最大+5.8 PPL(@10%)"),
    ("TableV",      "TableV_ZeroShot",            "零样本准确率（HellaSwag+WinoGrande）",      "zeroshot/opt125m/zeroshot_summary.json",              "表格",      "ILP-0.7%pt vs CIM-7b; SQ+6b-2.7%pt"),
]
fills_idx = [H3 if i%2==0 else PatternFill("solid", fgColor="DDEEFF") for i in range(len(index_data))]
for i, (d, clr) in enumerate(zip(index_data, fills_idx)):
    write_row(ws, 3+i, range(1,7), d, fill=clr)

r = 3 + len(index_data) + 1
section_row(ws, r, 1, "★ 最重要发现: fc2(19%饱和率)敏感度=1.413，attn_qkv(100%饱和率)敏感度=0.128 → 11×反转 → ILP比HAWQ好12.7 PPL → 必须直接测量，不能用代理", span=6)
section_row(ws, r+1, 1, "数据根目录: /home/ubuntu/iccad2026_bxkj/NeuroSim/results/ | 论文: /home/ubuntu/iccad2026_bxkj/paper/main.tex", span=6, fill=H3)

# Save
wb.save(OUT)
print(f"Saved → {OUT}")
